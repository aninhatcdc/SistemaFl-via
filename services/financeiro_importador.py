import re
import unicodedata
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


MESES = {
    "JAN": 1,
    "JANEIRO": 1,
    "FEV": 2,
    "FEVEREIRO": 2,
    "MAR": 3,
    "MARCO": 3,
    "ABR": 4,
    "ABRIL": 4,
    "MAI": 5,
    "MAIO": 5,
    "JUN": 6,
    "JUNHO": 6,
    "JUL": 7,
    "JULHO": 7,
    "AGO": 8,
    "AGOSTO": 8,
    "SET": 9,
    "SETEMBRO": 9,
    "OUT": 10,
    "OUTUBRO": 10,
    "NOV": 11,
    "NOVEMBRO": 11,
    "DEZ": 12,
    "DEZEMBRO": 12
}


def normalizar_texto(valor):
    """
    Remove acentos, espaços duplicados e converte para maiúsculas.
    """

    if valor is None:
        return ""

    texto = str(valor).strip()

    texto = unicodedata.normalize(
        "NFKD",
        texto
    ).encode(
        "ASCII",
        "ignore"
    ).decode(
        "ASCII"
    )

    texto = re.sub(
        r"\s+",
        " ",
        texto
    )

    return texto.upper()


def limpar_descricao(valor):
    """
    Mantém os acentos da descrição, mas remove espaços extras.
    """

    if valor is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(valor).strip()
    )


def identificar_ano(nome_arquivo, ano_padrao=None):
    """
    Procura um ano de quatro dígitos no nome do arquivo.
    Exemplo: PLANILHA DE GASTOS - 2026.xlsx
    """

    nome_arquivo = nome_arquivo or ""

    resultado = re.search(
        r"\b(20\d{2})\b",
        nome_arquivo
    )

    if resultado:
        return int(resultado.group(1))

    if ano_padrao:
        return int(ano_padrao)

    return date.today().year


def converter_valor(valor):
    """
    Converte números e textos monetários para Decimal.

    Exemplos aceitos:
    1500
    1500.50
    1.500,50
    R$ 1.500,50
    R$96,98
    -

    Retorna:
    - Decimal quando o valor é válido;
    - None quando a célula está vazia ou contém "-";
    - levanta ValueError quando existe texto não reconhecido.
    """

    if valor is None:
        return None

    if isinstance(valor, bool):
        raise ValueError("Valor booleano não é aceito.")

    if isinstance(valor, (int, float, Decimal)):
        return Decimal(
            str(valor)
        ).quantize(
            Decimal("0.01")
        )

    texto_original = str(valor).strip()

    if not texto_original:
        return None

    if texto_original in {"-", "–", "—"}:
        return None

    # Remove caracteres invisíveis encontrados em algumas células do Excel.
    texto = (
        texto_original
        .replace("\u202c", "")
        .replace("\u202a", "")
        .replace("\u200e", "")
        .replace("\u200f", "")
        .replace("\xa0", "")
        .replace("R$", "")
        .replace("r$", "")
        .replace(" ", "")
    )

    # Mantém apenas números, vírgula, ponto e sinal.
    texto = re.sub(
        r"[^0-9,.\-]",
        "",
        texto
    )

    if not texto:
        raise ValueError(
            f"Conteúdo não numérico: {texto_original}"
        )

    # Formato brasileiro: 1.545,20
    if "," in texto:
        texto = texto.replace(".", "")
        texto = texto.replace(",", ".")

    # Caso haja vários pontos, considera o último como decimal.
    elif texto.count(".") > 1:
        partes = texto.split(".")
        texto = "".join(partes[:-1]) + "." + partes[-1]

    try:
        numero = Decimal(texto).quantize(
            Decimal("0.01")
        )
    except InvalidOperation as erro:
        raise ValueError(
            f"Valor inválido: {texto_original}"
        ) from erro

    if numero < 0:
        raise ValueError(
            f"Valor negativo não permitido: {texto_original}"
        )

    return numero


def localizar_cabecalho(planilha):
    """
    Procura uma linha cuja primeira célula seja ITENS, DESPESA ou DESCRIÇÃO.
    """

    termos_validos = {
        "ITENS",
        "ITEM",
        "DESPESA",
        "DESPESAS",
        "DESCRICAO"
    }

    limite = min(
        planilha.max_row,
        20
    )

    for numero_linha in range(1, limite + 1):
        primeira_celula = normalizar_texto(
            planilha.cell(
                row=numero_linha,
                column=1
            ).value
        )

        if primeira_celula in termos_validos:
            return numero_linha

    raise ValueError(
        "Não foi possível localizar o cabeçalho da planilha. "
        "A primeira coluna deve possuir o título ITENS ou DESPESA."
    )


def localizar_colunas_meses(planilha, linha_cabecalho):
    """
    Retorna um dicionário no formato:
    {numero_da_coluna: numero_do_mes}
    """

    colunas_meses = {}

    for coluna in range(
        2,
        planilha.max_column + 1
    ):
        cabecalho_original = planilha.cell(
            row=linha_cabecalho,
            column=coluna
        ).value

        cabecalho = normalizar_texto(
            cabecalho_original
        )

        # A planilha atual possui EXT. DEZ.
        # Essa coluna não representa um mês comum e será ignorada.
        if cabecalho.startswith("EXT"):
            continue

        numero_mes = MESES.get(cabecalho)

        if numero_mes:
            colunas_meses[coluna] = numero_mes

    if not colunas_meses:
        raise ValueError(
            "Nenhuma coluna de mês foi encontrada na planilha."
        )

    return colunas_meses


def definir_status(ano, mes):
    """
    Meses futuros são importados como Previsto.
    Meses passados ou atuais são importados como Pago.
    """

    hoje = date.today()

    if ano > hoje.year:
        return "Previsto"

    if ano == hoje.year and mes > hoje.month:
        return "Previsto"

    return "Pago"


def ler_planilha_financeira(
    arquivo,
    nome_arquivo="",
    ano_padrao=None
):
    """
    Lê a planilha e devolve uma prévia dos lançamentos.

    O parâmetro arquivo pode ser:
    - caminho de arquivo;
    - stream recebido pelo Flask;
    - FileStorage.stream.
    """

    ano = identificar_ano(
        nome_arquivo,
        ano_padrao
    )

    workbook = load_workbook(
        arquivo,
        data_only=True,
        read_only=True
    )

    planilha = workbook.active

    linha_cabecalho = localizar_cabecalho(
        planilha
    )

    colunas_meses = localizar_colunas_meses(
        planilha,
        linha_cabecalho
    )

    lancamentos = []
    avisos = []
    itens_encontrados = set()

    primeira_linha_dados = linha_cabecalho + 1

    for numero_linha in range(
        primeira_linha_dados,
        planilha.max_row + 1
    ):
        descricao = limpar_descricao(
            planilha.cell(
                row=numero_linha,
                column=1
            ).value
        )

        if not descricao:
            continue

        descricao_normalizada = normalizar_texto(
            descricao
        )

        # Ignora linhas de totalização caso existam.
        if descricao_normalizada.startswith("TOTAL"):
            continue

        valores_validos_item = 0

        for coluna, mes in colunas_meses.items():
            conteudo = planilha.cell(
                row=numero_linha,
                column=coluna
            ).value

            try:
                valor = converter_valor(
                    conteudo
                )
            except ValueError:
                avisos.append({
                    "linha": numero_linha,
                    "mes": mes,
                    "descricao": descricao,
                    "conteudo": str(conteudo),
                    "mensagem": (
                        "A célula contém um texto que não pôde "
                        "ser convertido em valor."
                    )
                })

                continue

            if valor is None:
                continue

            if valor == Decimal("0.00"):
                continue

            valores_validos_item += 1
            itens_encontrados.add(descricao)

            lancamentos.append({
                "tipo": "Despesa",
                "descricao": descricao,
                "categoria": "Outros",
                "valor": valor,
                "competencia_mes": mes,
                "competencia_ano": ano,
                "status": definir_status(
                    ano,
                    mes
                ),
                "recorrente": False,
                "observacoes": (
                    f"Importado da planilha {nome_arquivo}"
                    if nome_arquivo
                    else "Importado de planilha Excel"
                )
            })

        # Considera recorrente quando aparece em pelo menos três meses.
        if valores_validos_item >= 3:
            for lancamento in lancamentos:
                if lancamento["descricao"] == descricao:
                    lancamento["recorrente"] = True

    workbook.close()

    totais_meses = {
        mes: Decimal("0.00")
        for mes in range(1, 13)
    }

    for lancamento in lancamentos:
        totais_meses[
            lancamento["competencia_mes"]
        ] += lancamento["valor"]

    total_geral = sum(
        totais_meses.values(),
        Decimal("0.00")
    )

    return {
        "ano": ano,
        "nome_planilha": planilha.title,
        "quantidade_itens": len(itens_encontrados),
        "quantidade_lancamentos": len(lancamentos),
        "lancamentos": lancamentos,
        "avisos": avisos,
        "totais_meses": totais_meses,
        "total_geral": total_geral
    }