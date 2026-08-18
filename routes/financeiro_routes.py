import os

from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import uuid4

from flask import (
    Blueprint,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for
)

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side
)

from werkzeug.utils import secure_filename

from models import db
from models.financeiro import LancamentoFinanceiro
from services.financeiro_importador import ler_planilha_financeira
from utils.permissoes import financeiro_required


financeiro_bp = Blueprint(
    "financeiro",
    __name__
)


MESES = [
    (1, "Janeiro", "JAN"),
    (2, "Fevereiro", "FEV"),
    (3, "Março", "MAR"),
    (4, "Abril", "ABR"),
    (5, "Maio", "MAI"),
    (6, "Junho", "JUN"),
    (7, "Julho", "JUL"),
    (8, "Agosto", "AGO"),
    (9, "Setembro", "SET"),
    (10, "Outubro", "OUT"),
    (11, "Novembro", "NOV"),
    (12, "Dezembro", "DEZ")
]


EXTENSOES_EXCEL = {
    "xlsx"
}


# =====================================
# FUNÇÕES AUXILIARES
# =====================================
def extensao_permitida(nome_arquivo):
    return (
        "." in nome_arquivo
        and nome_arquivo.rsplit(
            ".",
            1
        )[1].lower() in EXTENSOES_EXCEL
    )


def pasta_importacoes():
    pasta = os.path.join(
        current_app.instance_path,
        "importacoes_financeiro"
    )

    os.makedirs(
        pasta,
        exist_ok=True
    )

    return pasta


def caminho_temporario_seguro(token):
    token_seguro = os.path.basename(
        str(
            token or ""
        )
    )

    if not token_seguro:
        return None

    caminho = os.path.join(
        pasta_importacoes(),
        token_seguro
    )

    if not os.path.isfile(
        caminho
    ):
        return None

    return caminho


def converter_valor_br(valor_texto):
    """
    Converte valores como:

    1250,50
    1.250,50
    R$ 1.250,50
    """

    if valor_texto is None:
        return None

    valor_texto = str(
        valor_texto
    ).strip()

    if not valor_texto:
        return None

    valor_texto = (
        valor_texto
        .replace(
            "R$",
            ""
        )
        .replace(
            " ",
            ""
        )
        .replace(
            ".",
            ""
        )
        .replace(
            ",",
            "."
        )
    )

    try:
        valor = Decimal(
            valor_texto
        )

        if valor < 0:
            raise ValueError(
                "O valor não pode ser negativo."
            )

        return valor.quantize(
            Decimal(
                "0.01"
            )
        )

    except (
        InvalidOperation,
        ValueError
    ):
        raise ValueError(
            "Digite um valor válido."
        )


def calcular_totais(
    ano,
    descricao
):
    lancamentos_ano = (
        LancamentoFinanceiro.query
        .filter_by(
            competencia_ano=ano
        )
        .all()
    )

    totais_meses = {
        numero: Decimal(
            "0.00"
        )
        for numero, _, _ in MESES
    }

    total_linha = Decimal(
        "0.00"
    )

    for lancamento in lancamentos_ano:
        valor = Decimal(
            lancamento.valor or 0
        )

        totais_meses[
            lancamento.competencia_mes
        ] += valor

        if (
            lancamento.descricao.strip()
            == descricao.strip()
        ):
            total_linha += valor

    total_anual = sum(
        totais_meses.values(),
        Decimal(
            "0.00"
        )
    )

    media_mensal = (
        total_anual / Decimal(
            "12"
        )
        if total_anual
        else Decimal(
            "0.00"
        )
    )

    return {
        "totais_meses": {
            str(numero): float(
                valor
            )
            for numero, valor
            in totais_meses.items()
        },
        "total_linha": float(
            total_linha
        ),
        "total_anual": float(
            total_anual
        ),
        "media_mensal": float(
            media_mensal
        )
    }


# =====================================
# CENTRAL FINANCEIRA
# =====================================
@financeiro_bp.route(
    "/financeiro"
)
@financeiro_required
def listar_financeiro():
    hoje = date.today()

    ano_atual = hoje.year
    mes_atual = hoje.month

    ano_selecionado = request.args.get(
        "ano",
        ano_atual,
        type=int
    )

    termo = request.args.get(
        "q",
        ""
    ).strip()

    consulta = (
        LancamentoFinanceiro.query
        .filter_by(
            competencia_ano=ano_selecionado
        )
    )

    if termo:
        consulta = consulta.filter(
            LancamentoFinanceiro.descricao.contains(
                termo
            )
            |
            LancamentoFinanceiro.categoria.contains(
                termo
            )
        )

    lancamentos = (
        consulta
        .order_by(
            LancamentoFinanceiro.descricao.asc(),
            LancamentoFinanceiro.competencia_mes.asc()
        )
        .all()
    )

    linhas = {}

    totais_meses = {
        numero: Decimal(
            "0.00"
        )
        for numero, _, _ in MESES
    }

    for lancamento in lancamentos:
        chave = (
            lancamento
            .descricao
            .strip()
        )

        if chave not in linhas:
            linhas[chave] = {
                "descricao": lancamento.descricao,
                "categoria": lancamento.categoria or "",
                "valores": {
                    numero: {
                        "id": None,
                        "valor": Decimal(
                            "0.00"
                        ),
                        "status": "",
                        "tipo": "",
                        "observacoes": ""
                    }
                    for numero, _, _ in MESES
                }
            }

        valor = Decimal(
            lancamento.valor or 0
        )

        linhas[chave]["valores"][
            lancamento.competencia_mes
        ] = {
            "id": lancamento.id,
            "valor": valor,
            "status": lancamento.status or "",
            "tipo": lancamento.tipo or "",
            "observacoes": (
                lancamento.observacoes
                or ""
            )
        }

        totais_meses[
            lancamento.competencia_mes
        ] += valor

    linhas = list(
        linhas.values()
    )

    total_anual = sum(
        totais_meses.values(),
        Decimal(
            "0.00"
        )
    )

    media_mensal = (
        total_anual / Decimal(
            "12"
        )
        if total_anual
        else Decimal(
            "0.00"
        )
    )

    anos_banco = (
        LancamentoFinanceiro.query
        .with_entities(
            LancamentoFinanceiro.competencia_ano
        )
        .distinct()
        .order_by(
            LancamentoFinanceiro.competencia_ano.desc()
        )
        .all()
    )

    anos_disponiveis = {
        ano_atual,
        ano_selecionado
    }

    for resultado in anos_banco:
        anos_disponiveis.add(
            resultado[0]
        )

    anos_disponiveis = sorted(
        anos_disponiveis,
        reverse=True
    )

    return render_template(
        "financeiro/listar.html",
        meses=MESES,
        linhas=linhas,
        totais_meses=totais_meses,
        total_anual=total_anual,
        media_mensal=media_mensal,
        ano_selecionado=ano_selecionado,
        anos_disponiveis=anos_disponiveis,
        termo=termo,
        mes_atual=mes_atual,
        ano_atual=ano_atual
    )


# =====================================
# NOVA DESPESA
# =====================================
@financeiro_bp.route(
    "/financeiro/despesas/nova",
    methods=[
        "POST"
    ]
)
@financeiro_required
def nova_despesa():
    descricao = request.form.get(
        "descricao",
        ""
    ).strip()

    categoria = request.form.get(
        "categoria",
        ""
    ).strip()

    valor_texto = request.form.get(
        "valor",
        ""
    ).strip()

    mes = request.form.get(
        "mes",
        type=int
    )

    ano = request.form.get(
        "ano",
        type=int
    )

    status = request.form.get(
        "status",
        "Pago"
    ).strip()

    observacoes = request.form.get(
        "observacoes",
        ""
    ).strip()

    recorrente = (
        request.form.get(
            "recorrente"
        )
        == "on"
    )

    if not descricao:
        flash(
            "Informe a descrição da despesa.",
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.listar_financeiro",
                ano=(
                    ano
                    or date.today().year
                )
            )
        )

    if (
        not mes
        or mes < 1
        or mes > 12
    ):
        flash(
            "Selecione um mês válido.",
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.listar_financeiro",
                ano=(
                    ano
                    or date.today().year
                )
            )
        )

    if (
        not ano
        or ano < 2000
        or ano > 2100
    ):
        flash(
            "Informe um ano válido.",
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.listar_financeiro"
            )
        )

    try:
        valor = converter_valor_br(
            valor_texto
        )

    except ValueError as erro:
        flash(
            str(
                erro
            ),
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.listar_financeiro",
                ano=ano
            )
        )

    if (
        valor is None
        or valor <= 0
    ):
        flash(
            "Informe um valor maior que zero.",
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.listar_financeiro",
                ano=ano
            )
        )

    status_validos = {
        "Pago",
        "Pendente",
        "Previsto"
    }

    if status not in status_validos:
        status = "Pago"

    despesa_existente = (
        LancamentoFinanceiro.query
        .filter_by(
            descricao=descricao,
            competencia_mes=mes,
            competencia_ano=ano
        )
        .first()
    )

    if despesa_existente:
        flash(
            (
                "Já existe uma despesa com essa descrição "
                "no mês e ano selecionados. "
                "Edite o valor diretamente na célula."
            ),
            "warning"
        )

        return redirect(
            url_for(
                "financeiro.listar_financeiro",
                ano=ano
            )
        )

    despesa = LancamentoFinanceiro(
        tipo="Despesa",
        descricao=descricao,
        categoria=(
            categoria
            or "Outros"
        ),
        valor=valor,
        competencia_mes=mes,
        competencia_ano=ano,
        status=status,
        recorrente=recorrente,
        observacoes=observacoes
    )

    db.session.add(
        despesa
    )

    db.session.commit()

    flash(
        "✅ Nova despesa cadastrada com sucesso!",
        "success"
    )

    return redirect(
        url_for(
            "financeiro.listar_financeiro",
            ano=ano
        )
    )


# =====================================
# EXCLUIR DESPESA DO ANO
# =====================================
@financeiro_bp.route(
    "/financeiro/despesas/excluir",
    methods=[
        "POST"
    ]
)
@financeiro_required
def excluir_despesa():
    descricao = request.form.get(
        "descricao",
        ""
    ).strip()

    ano = request.form.get(
        "ano",
        type=int
    )

    if (
        not descricao
        or not ano
    ):
        flash(
            "Não foi possível identificar a despesa.",
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.listar_financeiro"
            )
        )

    lancamentos = (
        LancamentoFinanceiro.query
        .filter_by(
            descricao=descricao,
            competencia_ano=ano
        )
        .all()
    )

    if not lancamentos:
        flash(
            "Nenhum lançamento dessa despesa foi encontrado.",
            "warning"
        )

        return redirect(
            url_for(
                "financeiro.listar_financeiro",
                ano=ano
            )
        )

    quantidade = len(
        lancamentos
    )

    for lancamento in lancamentos:
        db.session.delete(
            lancamento
        )

    db.session.commit()

    flash(
        (
            "🗑 Despesa removida com sucesso. "
            f"{quantidade} lançamento(s) foram excluídos."
        ),
        "success"
    )

    return redirect(
        url_for(
            "financeiro.listar_financeiro",
            ano=ano
        )
    )


# =====================================
# IMPORTAR PLANILHA — SELEÇÃO E PRÉVIA
# =====================================
@financeiro_bp.route(
    "/financeiro/importar",
    methods=[
        "GET",
        "POST"
    ]
)
@financeiro_required
def importar_excel():
    if request.method == "GET":
        return render_template(
            "financeiro/importar.html",
            previa=None,
            token=None,
            nome_arquivo=None
        )

    arquivo = request.files.get(
        "arquivo"
    )

    ano_informado = request.form.get(
        "ano",
        type=int
    )

    if (
        not arquivo
        or not arquivo.filename
    ):
        flash(
            "Selecione uma planilha Excel.",
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.importar_excel"
            )
        )

    if not extensao_permitida(
        arquivo.filename
    ):
        flash(
            "Envie um arquivo no formato .xlsx.",
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.importar_excel"
            )
        )

    nome_original = secure_filename(
        arquivo.filename
    )

    token = (
        f"{uuid4().hex}_"
        f"{nome_original}"
    )

    caminho = os.path.join(
        pasta_importacoes(),
        token
    )

    arquivo.save(
        caminho
    )

    try:
        previa = ler_planilha_financeira(
            caminho,
            nome_arquivo=nome_original,
            ano_padrao=ano_informado
        )

    except Exception as erro:
        if os.path.exists(
            caminho
        ):
            os.remove(
                caminho
            )

        flash(
            (
                "Não foi possível ler a planilha: "
                f"{erro}"
            ),
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.importar_excel"
            )
        )

    return render_template(
        "financeiro/importar.html",
        previa=previa,
        token=token,
        nome_arquivo=nome_original
    )


# =====================================
# CONFIRMAR IMPORTAÇÃO
# =====================================
@financeiro_bp.route(
    "/financeiro/importar/confirmar",
    methods=[
        "POST"
    ]
)
@financeiro_required
def confirmar_importacao():
    token = request.form.get(
        "token",
        ""
    )

    ano_confirmado = request.form.get(
        "ano",
        type=int
    )

    caminho = caminho_temporario_seguro(
        token
    )

    if not caminho:
        flash(
            (
                "O arquivo temporário não foi encontrado. "
                "Selecione a planilha novamente."
            ),
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.importar_excel"
            )
        )

    nome_original = token.split(
        "_",
        1
    )[-1]

    dados = None

    try:
        dados = ler_planilha_financeira(
            caminho,
            nome_arquivo=nome_original,
            ano_padrao=ano_confirmado
        )

        criados = 0
        atualizados = 0

        for item in dados[
            "lancamentos"
        ]:
            lancamento = (
                LancamentoFinanceiro.query
                .filter_by(
                    descricao=item[
                        "descricao"
                    ],
                    competencia_mes=item[
                        "competencia_mes"
                    ],
                    competencia_ano=item[
                        "competencia_ano"
                    ]
                )
                .first()
            )

            if lancamento:
                lancamento.tipo = item[
                    "tipo"
                ]

                lancamento.categoria = item[
                    "categoria"
                ]

                lancamento.valor = item[
                    "valor"
                ]

                lancamento.status = item[
                    "status"
                ]

                lancamento.recorrente = item[
                    "recorrente"
                ]

                lancamento.observacoes = item[
                    "observacoes"
                ]

                atualizados += 1

            else:
                lancamento = LancamentoFinanceiro(
                    tipo=item[
                        "tipo"
                    ],
                    descricao=item[
                        "descricao"
                    ],
                    categoria=item[
                        "categoria"
                    ],
                    valor=item[
                        "valor"
                    ],
                    competencia_mes=item[
                        "competencia_mes"
                    ],
                    competencia_ano=item[
                        "competencia_ano"
                    ],
                    status=item[
                        "status"
                    ],
                    recorrente=item[
                        "recorrente"
                    ],
                    observacoes=item[
                        "observacoes"
                    ]
                )

                db.session.add(
                    lancamento
                )

                criados += 1

        db.session.commit()

    except Exception as erro:
        db.session.rollback()

        flash(
            (
                "Erro ao importar a planilha: "
                f"{erro}"
            ),
            "danger"
        )

        return redirect(
            url_for(
                "financeiro.importar_excel"
            )
        )

    finally:
        if os.path.exists(
            caminho
        ):
            os.remove(
                caminho
            )

    flash(
        (
            "✅ Importação concluída: "
            f"{criados} lançamentos criados e "
            f"{atualizados} atualizados."
        ),
        "success"
    )

    if (
        dados
        and dados["avisos"]
    ):
        flash(
            (
                f"⚠️ {len(dados['avisos'])} célula(s) "
                "não foram importadas por conterem "
                "valores inválidos."
            ),
            "warning"
        )

    return redirect(
        url_for(
            "financeiro.listar_financeiro",
            ano=dados["ano"]
        )
    )


# =====================================
# GRÁFICOS E COMPARATIVO FINANCEIRO
# =====================================
@financeiro_bp.route(
    "/financeiro/graficos"
)
@financeiro_required
def graficos_financeiros():
    hoje = date.today()

    anos_banco = (
        LancamentoFinanceiro.query
        .with_entities(
            LancamentoFinanceiro.competencia_ano
        )
        .distinct()
        .order_by(
            LancamentoFinanceiro.competencia_ano.desc()
        )
        .all()
    )

    anos_disponiveis = {
        hoje.year,
        hoje.year - 1
    }

    for resultado in anos_banco:
        if resultado[0]:
            anos_disponiveis.add(
                resultado[0]
            )

    anos_disponiveis = sorted(
        anos_disponiveis,
        reverse=True
    )

    ano_a = request.args.get(
        "ano_a",
        hoje.year - 1,
        type=int
    )

    ano_b = request.args.get(
        "ano_b",
        hoje.year,
        type=int
    )

    anos_disponiveis = sorted(
        (
            set(
                anos_disponiveis
            )
            |
            {
                ano_a,
                ano_b
            }
        ),
        reverse=True
    )

    def obter_dados_ano(ano):
        lancamentos = (
            LancamentoFinanceiro.query
            .filter_by(
                competencia_ano=ano
            )
            .all()
        )

        totais_meses = {
            numero: Decimal(
                "0.00"
            )
            for numero, _, _ in MESES
        }

        totais_categorias = {}

        for lancamento in lancamentos:
            valor = Decimal(
                lancamento.valor or 0
            )

            mes = (
                lancamento
                .competencia_mes
            )

            categoria = (
                lancamento.categoria
                or "Outros"
            )

            if mes in totais_meses:
                totais_meses[
                    mes
                ] += valor

            totais_categorias[
                categoria
            ] = (
                totais_categorias.get(
                    categoria,
                    Decimal(
                        "0.00"
                    )
                )
                + valor
            )

        total_anual = sum(
            totais_meses.values(),
            Decimal(
                "0.00"
            )
        )

        meses_com_valor = sum(
            1
            for valor
            in totais_meses.values()
            if valor > 0
        )

        media_mensal = (
            total_anual
            / Decimal(
                meses_com_valor
            )
            if meses_com_valor
            else Decimal(
                "0.00"
            )
        )

        maior_mes_numero = max(
            totais_meses,
            key=totais_meses.get
        )

        maior_mes_valor = totais_meses[
            maior_mes_numero
        ]

        maior_mes_nome = next(
            nome
            for numero, nome, _
            in MESES
            if numero == maior_mes_numero
        )

        return {
            "totais_meses": totais_meses,
            "totais_categorias": totais_categorias,
            "total_anual": total_anual,
            "media_mensal": media_mensal,
            "maior_mes_nome": maior_mes_nome,
            "maior_mes_valor": maior_mes_valor
        }

    dados_a = obter_dados_ano(
        ano_a
    )

    dados_b = obter_dados_ano(
        ano_b
    )

    diferenca = (
        dados_b["total_anual"]
        - dados_a["total_anual"]
    )

    variacao_percentual = None

    if dados_a[
        "total_anual"
    ] > 0:
        variacao_percentual = (
            diferenca
            / dados_a["total_anual"]
            * Decimal(
                "100"
            )
        )

    categorias_ordenadas = sorted(
        dados_b[
            "totais_categorias"
        ].items(),
        key=lambda item: item[1],
        reverse=True
    )

    return render_template(
        "financeiro/graficos.html",
        meses=MESES,
        anos_disponiveis=anos_disponiveis,
        ano_a=ano_a,
        ano_b=ano_b,
        dados_a=dados_a,
        dados_b=dados_b,
        diferenca=diferenca,
        variacao_percentual=variacao_percentual,
        categorias_ordenadas=categorias_ordenadas,
        rotulos_meses=[
            sigla
            for _, _, sigla
            in MESES
        ],
        valores_ano_a=[
            float(
                dados_a[
                    "totais_meses"
                ][numero]
            )
            for numero, _, _
            in MESES
        ],
        valores_ano_b=[
            float(
                dados_b[
                    "totais_meses"
                ][numero]
            )
            for numero, _, _
            in MESES
        ],
        categorias_rotulos=[
            categoria
            for categoria, _
            in categorias_ordenadas
        ],
        categorias_valores=[
            float(
                valor
            )
            for _, valor
            in categorias_ordenadas
        ]
    )


# =====================================
# EXPORTAR PLANILHA EXCEL
# =====================================
@financeiro_bp.route(
    "/financeiro/exportar"
)
@financeiro_required
def exportar_excel():
    ano = request.args.get(
        "ano",
        date.today().year,
        type=int
    )

    lancamentos = (
        LancamentoFinanceiro.query
        .filter_by(
            competencia_ano=ano
        )
        .order_by(
            LancamentoFinanceiro.descricao.asc(),
            LancamentoFinanceiro.competencia_mes.asc()
        )
        .all()
    )

    linhas = {}

    totais_meses = {
        numero: Decimal(
            "0.00"
        )
        for numero, _, _ in MESES
    }

    for lancamento in lancamentos:
        descricao = (
            lancamento
            .descricao
            .strip()
        )

        if descricao not in linhas:
            linhas[descricao] = {
                "categoria": (
                    lancamento.categoria
                    or "Outros"
                ),
                "valores": {
                    numero: Decimal(
                        "0.00"
                    )
                    for numero, _, _ in MESES
                }
            }

        valor = Decimal(
            lancamento.valor or 0
        )

        linhas[
            descricao
        ]["valores"][
            lancamento.competencia_mes
        ] += valor

        totais_meses[
            lancamento.competencia_mes
        ] += valor

    workbook = Workbook()

    planilha = workbook.active

    planilha.title = (
        f"Financeiro {ano}"
    )

    cabecalhos = [
        "Despesa",
        "Categoria"
    ]

    cabecalhos.extend(
        sigla
        for _, _, sigla
        in MESES
    )

    cabecalhos.append(
        "Total"
    )

    planilha.append(
        cabecalhos
    )

    for descricao, dados in linhas.items():
        total_linha = sum(
            dados[
                "valores"
            ].values(),
            Decimal(
                "0.00"
            )
        )

        linha_excel = [
            descricao,
            dados["categoria"]
        ]

        linha_excel.extend(
            float(
                dados[
                    "valores"
                ][numero]
            )
            for numero, _, _
            in MESES
        )

        linha_excel.append(
            float(
                total_linha
            )
        )

        planilha.append(
            linha_excel
        )

    total_anual = sum(
        totais_meses.values(),
        Decimal(
            "0.00"
        )
    )

    linha_total = [
        "TOTAL MENSAL",
        ""
    ]

    linha_total.extend(
        float(
            totais_meses[
                numero
            ]
        )
        for numero, _, _
        in MESES
    )

    linha_total.append(
        float(
            total_anual
        )
    )

    planilha.append(
        linha_total
    )

    preenchimento_cabecalho = PatternFill(
        fill_type="solid",
        fgColor="212529"
    )

    fonte_cabecalho = Font(
        color="FFFFFF",
        bold=True
    )

    preenchimento_total = PatternFill(
        fill_type="solid",
        fgColor="FFF4CC"
    )

    fonte_total = Font(
        color="0D6EFD",
        bold=True
    )

    preenchimento_total_anual = PatternFill(
        fill_type="solid",
        fgColor="198754"
    )

    fonte_total_anual = Font(
        color="FFFFFF",
        bold=True
    )

    borda_fina = Border(
        left=Side(
            style="thin",
            color="D9D9D9"
        ),
        right=Side(
            style="thin",
            color="D9D9D9"
        ),
        top=Side(
            style="thin",
            color="D9D9D9"
        ),
        bottom=Side(
            style="thin",
            color="D9D9D9"
        )
    )

    for celula in planilha[1]:
        celula.fill = (
            preenchimento_cabecalho
        )

        celula.font = (
            fonte_cabecalho
        )

        celula.alignment = Alignment(
            horizontal="center"
        )

        celula.border = borda_fina

    ultima_linha = (
        planilha.max_row
    )

    ultima_coluna = (
        planilha.max_column
    )

    for celula in planilha[
        ultima_linha
    ]:
        celula.fill = (
            preenchimento_total
        )

        celula.font = fonte_total

        celula.border = borda_fina

    celula_total_anual = planilha.cell(
        row=ultima_linha,
        column=ultima_coluna
    )

    celula_total_anual.fill = (
        preenchimento_total_anual
    )

    celula_total_anual.font = (
        fonte_total_anual
    )

    for linha in planilha.iter_rows(
        min_row=2,
        max_row=planilha.max_row,
        min_col=3,
        max_col=planilha.max_column
    ):
        for celula in linha:
            celula.number_format = (
                'R$ #,##0.00'
            )

            celula.alignment = Alignment(
                horizontal="right"
            )

            celula.border = (
                borda_fina
            )

    for linha in planilha.iter_rows(
        min_row=2,
        max_row=planilha.max_row,
        min_col=1,
        max_col=2
    ):
        for celula in linha:
            celula.border = (
                borda_fina
            )

    planilha.freeze_panes = (
        "C2"
    )

    planilha.auto_filter.ref = (
        f"A1:"
        f"{planilha.cell(
            row=1,
            column=planilha.max_column
        ).coordinate}"
    )

    planilha.column_dimensions[
        "A"
    ].width = 32

    planilha.column_dimensions[
        "B"
    ].width = 18

    for coluna in range(
        3,
        planilha.max_column + 1
    ):
        letra = planilha.cell(
            row=1,
            column=coluna
        ).column_letter

        planilha.column_dimensions[
            letra
        ].width = 14

    arquivo_memoria = BytesIO()

    workbook.save(
        arquivo_memoria
    )

    arquivo_memoria.seek(
        0
    )

    return send_file(
        arquivo_memoria,
        as_attachment=True,
        download_name=(
            f"financeiro_{ano}.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


# =====================================
# SALVAR CÉLULA DA PLANILHA
# =====================================
@financeiro_bp.route(
    "/financeiro/celula/salvar",
    methods=[
        "POST"
    ]
)
@financeiro_required
def salvar_celula():
    dados = request.get_json(
        silent=True
    ) or {}

    descricao = str(
        dados.get(
            "descricao",
            ""
        )
    ).strip()

    categoria = str(
        dados.get(
            "categoria",
            ""
        )
    ).strip()

    mes = dados.get(
        "mes"
    )

    ano = dados.get(
        "ano"
    )

    valor_texto = dados.get(
        "valor"
    )

    if not descricao:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                "A descrição da despesa é obrigatória."
            )
        }), 400

    try:
        mes = int(
            mes
        )

        ano = int(
            ano
        )

    except (
        TypeError,
        ValueError
    ):
        return jsonify({
            "sucesso": False,
            "mensagem": (
                "Mês ou ano inválido."
            )
        }), 400

    if (
        mes < 1
        or mes > 12
    ):
        return jsonify({
            "sucesso": False,
            "mensagem": (
                "O mês informado é inválido."
            )
        }), 400

    try:
        valor = converter_valor_br(
            valor_texto
        )

    except ValueError as erro:
        return jsonify({
            "sucesso": False,
            "mensagem": str(
                erro
            )
        }), 400

    lancamento = (
        LancamentoFinanceiro.query
        .filter_by(
            descricao=descricao,
            competencia_mes=mes,
            competencia_ano=ano
        )
        .first()
    )

    if valor is None:
        if lancamento:
            db.session.delete(
                lancamento
            )

            db.session.commit()

        totais = calcular_totais(
            ano,
            descricao
        )

        return jsonify({
            "sucesso": True,
            "removido": True,
            "id": None,
            "valor": None,
            **totais
        })

    if lancamento:
        lancamento.valor = valor

        lancamento.categoria = (
            categoria
            or lancamento.categoria
        )

    else:
        lancamento = LancamentoFinanceiro(
            tipo="Despesa",
            descricao=descricao,
            categoria=(
                categoria
                or "Outros"
            ),
            valor=valor,
            competencia_mes=mes,
            competencia_ano=ano,
            status="Pago"
        )

        db.session.add(
            lancamento
        )

    db.session.commit()

    totais = calcular_totais(
        ano,
        descricao
    )

    return jsonify({
        "sucesso": True,
        "removido": False,
        "id": lancamento.id,
        "valor": float(
            lancamento.valor
        ),
        **totais
    })